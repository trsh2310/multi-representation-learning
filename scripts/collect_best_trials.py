from pathlib import Path
from typing import Dict, List, Optional, Tuple

import json
import pandas as pd
import yaml
import click
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def load_yaml(path: Path) -> Dict:
    with path.open("r") as f:
        return yaml.safe_load(f) or {}


def find_config_file(directory: Path) -> Optional[Path]:
    candidates = [p for p in directory.glob("*.yaml") if p.name != "best_trial.yaml"]
    return sorted(candidates)[0] if candidates else None


def count_trials_from_log(directory: Path) -> int:
    """Count unique trial IDs from Optuna log file."""
    log_candidates = sorted(directory.glob("*.log"))
    if not log_candidates:
        return 0

    trial_ids = set()
    for log_path in log_candidates:
        try:
            with log_path.open("r") as f:
                for line in f:
                    line = line.strip()
                    if not line or not line.startswith("{"):
                        continue
                    try:
                        record = json.loads(line)
                    except json.JSONDecodeError:
                        continue
                    if "trial_id" in record:
                        trial_ids.add(record["trial_id"])
        except OSError:
            continue
    return len(trial_ids)


def extract_dataset_model(config: Dict) -> Tuple[Optional[str], Optional[str]]:
    """Extract dataset name and model name from config."""
    dataset_cfg = config.get("dataset") or {}
    model_cfg = config.get("model") or {}

    dataset = dataset_cfg.get("name") if isinstance(dataset_cfg, dict) else None
    model = None
    if isinstance(model_cfg, dict):
        model = model_cfg.get("name") or model_cfg.get("_target_")
        if model and "." in model and not model_cfg.get("name"):
            model = model.split(".")[-1]
    return dataset, model


def process_directory(directory: Path) -> Optional[Dict[str, object]]:
    """Process a single experiment directory and extract metrics if completed."""
    best_trial_path = directory / "best_trial.yaml"
    if not best_trial_path.is_file():
        return None

    config_path = find_config_file(directory)
    if not config_path or not config_path.is_file():
        return None

    best_data = load_yaml(best_trial_path)
    config_data = load_yaml(config_path)

    dataset, model = extract_dataset_model(config_data)

    user_attrs = best_data.get("user_attrs") or {}
    val_metric = user_attrs.get("val/ndcg@10", best_data.get("value"))

    row: Dict[str, object] = {
        "experiment": directory.name,
        "dataset": dataset,
        "model": model,
        "trial_number": best_data.get("number"),
        "n_trials": count_trials_from_log(directory),
        "val_metric": val_metric,
    }

    for key, value in user_attrs.items():
        if key.startswith("val/"):
            row[f"val_{key[4:]}"] = value

    for key, value in (best_data.get("test_metrics") or {}).items():
        row[key] = value

    return row


def collect_best_trials(root: Path) -> pd.DataFrame:
    """Collect all best trials from completed experiments."""
    rows: List[Dict[str, object]] = []
    for entry in sorted(root.iterdir()):
        if not entry.is_dir():
            continue
        row = process_directory(entry)
        if row and row.get("dataset") and row.get("model"):
            rows.append(row)
    return pd.DataFrame(rows)


def format_ci(mean: Optional[float], ci_low: Optional[float], ci_high: Optional[float], decimals: int = 4) -> str:
    if mean is None or ci_low is None or ci_high is None:
        return ""
    delta = max(mean - ci_low, ci_high - mean)
    mean_r = round(float(mean), decimals)
    delta_r = round(float(delta), decimals)
    return f"{mean_r:.{decimals}f} ± {delta_r:.{decimals}f}"


def build_human_view(df: pd.DataFrame, decimals: int = 4) -> pd.DataFrame:
    human = df.copy()

    human["val_metric"] = human["val_metric"].map(
        lambda x: f"{float(x):.{decimals}f}" if x is not None else ""
    )

    human["ndcg@10"] = human.apply(
        lambda row: format_ci(
            row.get("test/ndcg@10"),
            row.get("test/ndcg@10_ci_low"),
            row.get("test/ndcg@10_ci_high"),
            decimals,
        ),
        axis=1,
    )
    human["hr@10"] = human.apply(
        lambda row: format_ci(
            row.get("test/recall@10"),
            row.get("test/recall@10_ci_low"),
            row.get("test/recall@10_ci_high"),
            decimals,
        ),
        axis=1,
    )
    human["cov@10"] = human.apply(
        lambda row: format_ci(
            row.get("test/coverage@10"),
            row.get("test/coverage@10_ci_low"),
            row.get("test/coverage@10_ci_high"),
            decimals,
        ),
        axis=1,
    )

    metric_columns = ["ndcg@10", "hr@10", "cov@10", "val_metric"]
    base = human[["dataset", "model"] + metric_columns].copy()
    base["dataset"] = base["dataset"].fillna("")
    base["model"] = base["model"].fillna("")

    tables = []
    for metric in metric_columns:
        table = base.pivot(index="dataset", columns="model", values=metric)
        tables.append(table)

    human_view = pd.concat(tables, axis=1, keys=metric_columns)
    human_view.index.name = None  # Ensure no index name is written
    human_view.columns.names = ["metric", "model"]
    return human_view


def format_human_excel(path: Path, sheet_name: str = "Sheet1") -> None:
    wb = load_workbook(path)
    ws = wb[sheet_name]

    ws.delete_rows(3)

    max_row = ws.max_row
    max_col = ws.max_column

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.cell(row=1, column=1, value="dataset/model")

    for row in (1, 2):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = center
            cell.font = Font(bold=True)

    for row in range(3, max_row + 1):
        ws.cell(row=row, column=1).alignment = left

    ws.freeze_panes = "B3"

    dataset_width = 12
    for row in range(3, max_row + 1):
        value = ws.cell(row=row, column=1).value
        if value:
            dataset_width = max(dataset_width, len(str(value)) + 2)
    ws.column_dimensions["A"].width = dataset_width

    for col in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    wb.save(path)


def collect_log_metrics(root: Path) -> pd.DataFrame:
    """Collect per-trial metrics and params from Optuna log files."""
    rows: List[Dict[str, object]] = []

    for entry in sorted(root.iterdir()):
        if not entry.is_dir():
            continue

        log_candidates = sorted(entry.glob("*.log"))
        if not log_candidates:
            continue

        config_path = find_config_file(entry)
        config_data = load_yaml(config_path) if config_path else {}
        dataset, model = extract_dataset_model(config_data)
        if not dataset or not model:
            parts = entry.name.split("_", 1)
            if len(parts) == 2:
                model = model or parts[0]
                dataset = dataset or parts[1]

        trials: Dict[int, Dict[str, object]] = {}
        log_path = log_candidates[0]

        try:
            with log_path.open("r") as f:
                for line in f:
                    line = line.strip()
                    if not line or not line.startswith("{"):
                        continue
                    try:
                        record = json.loads(line)
                    except json.JSONDecodeError:
                        continue

                    if "trial_id" not in record:
                        continue
                    trial_id = record["trial_id"]
                    trial = trials.setdefault(
                        trial_id,
                        {
                            "experiment": entry.name,
                            "dataset": dataset,
                            "model": model,
                            "trial_id": trial_id,
                        },
                    )

                    if "param_name" in record:
                        trial[f"param_{record['param_name']}"] = record.get("param_value_internal")

                    if "user_attr" in record:
                        user_attr = record.get("user_attr") or {}
                        for key, value in user_attr.items():
                            if key.startswith("val/"):
                                trial[f"val_{key[4:]}"] = value
                            elif key == "error":
                                trial["error"] = value

                    if "values" in record and record.get("values"):
                        values = record.get("values")
                        if isinstance(values, list) and values:
                            trial["objective_value"] = values[0]

                    if "state" in record:
                        trial["state"] = record.get("state")
        except OSError:
            continue

        for trial in sorted(trials.values(), key=lambda item: item.get("trial_id", -1)):
            rows.append(trial)

    return pd.DataFrame(rows)


@click.command(help="Aggregate best trials across Optuna output folders (depth 1).")
@click.argument("root", type=click.Path(exists=True, file_okay=False, path_type=Path))
@click.option(
    "--output-dir",
    type=click.Path(file_okay=False, path_type=Path),
    default=None,
    help="Where to save the summary files. Defaults to the root folder.",
)
def main(root: Path, output_dir: Optional[Path]):
    output_dir = output_dir or root
    best_df = collect_best_trials(root)

    if best_df.empty:
        click.echo("No best_trial.yaml and config YAML pairs were found.")
        return

    output_dir.mkdir(parents=True, exist_ok=True)

    raw_csv_path = output_dir / "best_trials_raw.csv"
    best_df.to_csv(raw_csv_path, index=False)

    human_df = build_human_view(best_df)
    human_xlsx_path = output_dir / "best_trials_human.xlsx"
    try:
        human_df.to_excel(human_xlsx_path, merge_cells=True)
        format_human_excel(human_xlsx_path)
    except ImportError as e:
        click.echo(f"Saved {len(human_df)} rows to {raw_csv_path}. Excel save failed: {e}.")
        return

    click.echo(f"Saved outputs: {raw_csv_path}, {human_xlsx_path}.")


if __name__ == "__main__":
    main()
